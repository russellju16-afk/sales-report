import json
import re
from pathlib import Path
from datetime import datetime, date
import openpyxl

SRC = Path('/Users/russell/Downloads/销售利润表-20251222111220.xlsx')
OUT = Path('data/latest.json')
OLD = Path('data/latest.json')

# Category mapping based on product name/spec keywords.
RICE_KEYS = ['大米','粳米','籼米','香米','稻花香','长粒香','米']
OIL_KEYS = ['油','脂']
FLOUR_KEYS = ['面粉','小麦粉']
GRAIN_KEYS = ['杂粮','玉米','小米','高粱','荞麦','藜麦','黑米','红豆','绿豆','黄豆','薏米','燕麦','芸豆']


def to_date_str(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    if val is None:
        return ''
    s = str(val).strip()
    if len(s) >= 10:
        return s[:10]
    return s


def to_float(val):
    if val is None or val == '':
        return 0.0
    try:
        return float(val)
    except Exception:
        return 0.0


def norm_text(val):
    return '' if val is None else str(val).strip()


def product_label(name, spec):
    name = norm_text(name)
    spec = norm_text(spec)
    if spec:
        return f"{name} | {spec}"
    return name


def category_of(name, spec):
    text = f"{norm_text(name)} {norm_text(spec)}"
    if any(k in text for k in FLOUR_KEYS):
        return '面粉'
    if any(k in text for k in GRAIN_KEYS):
        return '杂粮'
    # avoid classifying 酱油 as 食用油
    if '酱油' not in text and any(k in text for k in OIL_KEYS):
        return '食用油'
    if any(k in text for k in RICE_KEYS):
        return '大米'
    return '其他'


def load_old_meta():
    if not OLD.exists():
        return {}, {}, {}, {}
    with OLD.open('r', encoding='utf-8') as f:
        old = json.load(f)
    data = old.get('data', {})
    return (
        data.get('order_map', {}),
        data.get('order_map_catton', {}),
        data.get('cat_ton', {}),
        data.get('cat_ton_meta', {"oil_density": 0.92, "fallback_bag_kg": 1, "missing_weight_lines": 0}),
    )


def main():
    if not SRC.exists():
        raise SystemExit(f"Source file not found: {SRC}")

    order_map_old, order_map_catton, cat_ton, cat_ton_meta = load_old_meta()

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cols = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    idx = {name: i for i, name in enumerate(cols)}

    segments = {
        'total': [],
        'store': [],
        'nonstore': [],
    }

    months = {
        'total': set(),
        'store': set(),
        'nonstore': set(),
    }

    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[0] is None:
            continue
        dt = to_date_str(r[idx['单据日期']])
        if not dt:
            continue
        month = dt[:7]
        order_no = norm_text(r[idx['单据编号']])
        cust = norm_text(r[idx['客户名称']])
        cls = norm_text(r[idx['客户分类']])
        name = norm_text(r[idx['商品名称']])
        spec = norm_text(r[idx['规格型号']])
        qty = to_float(r[idx['数量']])
        sales = to_float(r[idx['价税合计']])
        cost = to_float(r[idx['成本']])
        fee = to_float(r[idx.get('关联销售费用')])
        gp = to_float(r[idx.get('销售毛利')])
        if gp == 0 and (sales or cost):
            gp = sales - cost
        gp_adj = gp - fee
        unit_price = to_float(r[idx.get('实际含税单价')])
        cat = category_of(name, spec)
        prod_label = product_label(name, spec)

        row = [
            dt, order_no, cust, cls, name, spec, prod_label, cat,
            qty, sales, cost, fee, gp, gp_adj, unit_price
        ]

        segments['total'].append(row)
        months['total'].add(month)
        if cls == '超群门店':
            segments['store'].append(row)
            months['store'].add(month)
        else:
            segments['nonstore'].append(row)
            months['nonstore'].add(month)

    data = {
        'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'data': {
            'total': {'rows': segments['total'], 'months': sorted(months['total'])},
            'store': {'rows': segments['store'], 'months': sorted(months['store'])},
            'nonstore': {'rows': segments['nonstore'], 'months': sorted(months['nonstore'])},
            'order_map': {},
            'order_map_catton': order_map_catton,
            'cat_ton': cat_ton,
            'cat_ton_meta': cat_ton_meta,
        }
    }

    OUT.write_text(json.dumps(data, ensure_ascii=False), encoding='utf-8')
    print(f"Wrote {OUT} with {len(segments['total'])} rows")


if __name__ == '__main__':
    main()
